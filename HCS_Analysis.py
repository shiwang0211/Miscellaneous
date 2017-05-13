import openpyxl
import xml
import os
import xml.etree.ElementTree as ET

path=os.getcwd()
wb = openpyxl.load_workbook('testHCS.xlsx')

def Executable(TypeOfFacility):
    return {
        'Freeway': 'Freeways.exe',
        'Ramp': 'Ramps.exe',
        'Weave': 'Weaving.exe'
    }[TypeOfFacility]

def Extension(TypeOfFacility):
    return {
        'Freeway': '.xhf',
        'Ramp': '.xhr',
        'Weave': '.xhw'
    }[TypeOfFacility]

def getFileLocation(TypeOfFacility):
    return(path + '//HCS Analysis//' + TypeOfFacility + '_HCS')
    
def getFileName(TypeOfFacility, rowNum, InputOutput):
    return('Segment' + str(rowNum-1) + '_HCS_' + InputOutput + Extension(TypeOfFacility))

def getFileLocName(TypeOfFacility, rowNum, InputOutput):
    return(getFileLocation(TypeOfFacility) + '\\' + getFileName(rowNum, InputOutput, TypeOfFacility))

def RunHCS(TypeofFacility, Input, Output):
    os.system('h:')
    os.system('cd ' + getFileLocation(TypeOfFacility))
    os.system(Executable(TypeOfFacility) + ' /b ' + Input + ' ' + Output)

def HCS_Analysis(TypeOfFacility):

    tree = ET.parse(path + '//HCS Analysis//' + TypeOfFacility + 'Template.xml')
    root=tree.getroot()

    sheet = wb.get_sheet_by_name('HCS_' + TypeOfFacility)

    for rowNum in range(2, sheet.max_row + 1):
        for colNum in range(2, 4 + 1): #TBD if output is added
            attr_name = sheet.cell(row = 1,column = colNum).value
            attr_val = sheet.cell(row = rowNum,column = colNum).value
            for node in root.iter(attr_name):
                node.text = str(attr_val) 

        tree.write(getFileLocName(TypeOfFacility, rowNum, 'Input'))
        RunHCS(TypeOfFacility, getFileName(TypeOfFacility, rowNum, 'Input'), getFileName(TypeOfFacility, rowNum, 'Output'))
        rowNum += 1

    for rowNum in range(2, sheet.max_row + 1):
        tree = ET.parse(getFileLocName(TypeOfFacility,rowNum, 'Output'))
        root=tree.getroot()
        for colNum in range(6, sheet.max_column + 1): #TBD if output is added
            attr_name = sheet.cell(row = 1,column = colNum).value
            for node in root.iter(attr_name):
                attr_val = node.text

            sheet.cell(row = rowNum,column = colNum).value = attr_val

HCS_Analysis('Freeway')
#HCS_Analysis('Ramp')
#HCS_Analysis('Weave')
wb.save('updated_testHCS.xlsx')
